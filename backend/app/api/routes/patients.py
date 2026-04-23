from __future__ import annotations

import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import and_, select
from sqlalchemy.orm import Session

from backend.app.crud.patients import (
    create_patient,
    delete_patient,
    get_patient,
    list_patients,
    update_patient,
)
from backend.app.db.session import get_db
from backend.app.models.appointment import Consulta
from backend.app.schemas.history import HistoricoPacienteOut
from backend.app.schemas.patient import PacienteCreate, PacienteDetail, PacienteOut, PacienteUpdate
from backend.app.utils.deps import get_current_user
from backend.app.utils.pdf import gerar_historico_pdf

router = APIRouter()


@router.get("", response_model=list[PacienteOut])
def list_(
    search: str | None = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
) -> list[PacienteOut]:
    return list_patients(db, usuario_id=user.id, search=search)


@router.post("", response_model=PacienteOut, status_code=status.HTTP_201_CREATED)
def create(
    data: PacienteCreate,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
) -> PacienteOut:
    return create_patient(db, usuario_id=user.id, data=data)


@router.get("/{patient_id}", response_model=PacienteDetail)
def get(patient_id, db: Session = Depends(get_db), user=Depends(get_current_user)) -> PacienteDetail:
    patient = get_patient(db, patient_id)
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente não encontrado")
    if patient.usuario_id != user.id:
        raise HTTPException(status_code=404, detail="Paciente não encontrado")
    return patient


@router.put("/{patient_id}", response_model=PacienteDetail)
def update(
    patient_id,
    data: PacienteUpdate,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
) -> PacienteDetail:
    patient = get_patient(db, patient_id)
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente não encontrado")
    if patient.usuario_id != user.id:
        raise HTTPException(status_code=404, detail="Paciente não encontrado")
    return update_patient(db, patient, data)


@router.delete("/{patient_id}")
def delete(patient_id, db: Session = Depends(get_db), user=Depends(get_current_user)) -> dict:
    patient = get_patient(db, patient_id)
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente não encontrado")
    if patient.usuario_id != user.id:
        raise HTTPException(status_code=404, detail="Paciente não encontrado")
    try:
        delete_patient(db, patient)
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return {"ok": True}


@router.get("/{patient_id}/history", response_model=HistoricoPacienteOut)
def history(
    patient_id,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
) -> HistoricoPacienteOut:
    patient = get_patient(db, patient_id)
    if not patient or patient.usuario_id != user.id:
        raise HTTPException(status_code=404, detail="Paciente não encontrado")

    consultas = list(
        db.scalars(
            select(Consulta)
            .where(and_(Consulta.usuario_id == user.id, Consulta.paciente_id == patient.id))
            .order_by(Consulta.inicio.desc())
            .limit(50)
        ).all()
    )
    return HistoricoPacienteOut(paciente=patient, consultas=consultas)


@router.get("/{patient_id}/history/pdf")
def history_pdf(
    patient_id,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
) -> Response:
    patient = get_patient(db, patient_id)
    if not patient or patient.usuario_id != user.id:
        raise HTTPException(status_code=404, detail="Paciente não encontrado")

    consultas = list(
        db.scalars(
            select(Consulta)
            .where(and_(Consulta.usuario_id == user.id, Consulta.paciente_id == patient.id))
            .order_by(Consulta.inicio.desc())
            .limit(50)
        ).all()
    )

    pdf_bytes = gerar_historico_pdf(patient, consultas)
    nome_arquivo = f"historico_{patient.nome_completo.replace(' ', '_')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


# ─── Excel Export ─────────────────────────────────────────────────────────────

@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
) -> StreamingResponse:
    """Exporta todos os pacientes do usuário como planilha Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado.")

    pacientes = list_patients(db, usuario_id=user.id, search=None, limit=10_000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacientes"

    # Cabeçalho
    headers = ["Nome Completo", "Telefone", "E-mail", "Data Nascimento", "Observações"]
    header_fill = PatternFill("solid", fgColor="1A3A6B")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    # Dados
    for row_idx, p in enumerate(pacientes, 2):
        nasc = p.data_nascimento.strftime("%d/%m/%Y") if p.data_nascimento else ""
        ws.cell(row=row_idx, column=1, value=p.nome_completo)
        ws.cell(row=row_idx, column=2, value=p.telefone)
        ws.cell(row=row_idx, column=3, value=p.email or "")
        ws.cell(row=row_idx, column=4, value=nasc)
        ws.cell(row=row_idx, column=5, value=p.observacoes or "")

        # Zebra
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="EEF3FF")
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = fill

    # Largura das colunas
    col_widths = [40, 20, 35, 18, 50]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"pacientes.xlsx\""},
    )


# ─── Excel Import ─────────────────────────────────────────────────────────────

@router.post("/import/excel")
def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
) -> dict:
    """
    Importa pacientes de uma planilha Excel (.xlsx).

    Colunas esperadas (na primeira linha / cabeçalho):
      Nome Completo | Telefone | E-mail | Data Nascimento | Observações

    - Linhas com nome ou telefone em branco são ignoradas.
    - Pacientes com mesmo telefone já cadastrado são atualizados (upsert).
    - Data de nascimento aceita: DD/MM/YYYY ou YYYY-MM-DD.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx válido.")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado.")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo Excel inválido ou corrompido.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Planilha vazia.")

    # Detecta cabeçalho — mapeia nome da coluna → índice (case-insensitive)
    header_row = [str(c).strip().lower() if c else "" for c in rows[0]]
    col_map: dict[str, int] = {}
    aliases = {
        "nome completo": "nome",
        "nome": "nome",
        "telefone": "telefone",
        "fone": "telefone",
        "celular": "telefone",
        "e-mail": "email",
        "email": "email",
        "data nascimento": "nasc",
        "data de nascimento": "nasc",
        "nascimento": "nasc",
        "observações": "obs",
        "observacoes": "obs",
        "obs": "obs",
        "observação": "obs",
    }
    for idx, h in enumerate(header_row):
        key = aliases.get(h)
        if key and key not in col_map:
            col_map[key] = idx

    if "nome" not in col_map or "telefone" not in col_map:
        raise HTTPException(
            status_code=400,
            detail="A planilha precisa ter colunas 'Nome Completo' e 'Telefone'.",
        )

    def _parse_date(val) -> date | None:
        if not val:
            return None
        s = str(val).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return date.fromisoformat(s) if fmt == "%Y-%m-%d" else \
                    __import__("datetime").datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def _get(row, key) -> str:
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    # Carrega telefones existentes para upsert
    from sqlalchemy import select as sel_
    from backend.app.models.patient import Paciente as PacienteModel

    existing_by_tel: dict[str, PacienteModel] = {
        p.telefone: p
        for p in db.scalars(
            sel_(PacienteModel).where(PacienteModel.usuario_id == user.id)
        ).all()
    }

    criados = 0
    atualizados = 0
    ignorados = 0

    for row in rows[1:]:
        nome = _get(row, "nome")
        telefone = _get(row, "telefone")

        if not nome or not telefone:
            ignorados += 1
            continue

        email_val = _get(row, "email") or None
        nasc_val = _parse_date(_get(row, "nasc"))
        obs_val = _get(row, "obs") or None

        existing = existing_by_tel.get(telefone)
        if existing:
            # Atualiza apenas campos não-vazios
            existing.nome_completo = nome
            if email_val:
                existing.email = email_val
            if nasc_val:
                existing.data_nascimento = nasc_val
            if obs_val:
                existing.observacoes = obs_val
            db.add(existing)
            atualizados += 1
        else:
            data = PacienteCreate(
                nome_completo=nome,
                telefone=telefone,
                email=email_val,
                data_nascimento=nasc_val,
                observacoes=obs_val,
            )
            novo = create_patient(db, usuario_id=user.id, data=data)
            existing_by_tel[telefone] = novo
            criados += 1

    db.commit()

    return {
        "ok": True,
        "criados": criados,
        "atualizados": atualizados,
        "ignorados": ignorados,
        "total": criados + atualizados,
    }
